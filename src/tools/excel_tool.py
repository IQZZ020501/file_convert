import logging
from pathlib import Path
from typing import Dict, Any

import openpyxl

logger = logging.getLogger("file_to_markdown")


async def excel_to_markdown(arguments: Dict[str, Any]) -> str:
    """将Excel文件转换为Markdown格式"""
    try:
        file_path = arguments.get("file_path")
        sheet_name = arguments.get("sheet_name", None)

        if not sheet_name:
            sheet_name = "sheet1"

        if not file_path:
            return "错误: 需要提供Excel文件路径"

        excel_path = Path(file_path)
        if not excel_path.exists():
            return f"错误: 文件不存在: {file_path}"

        if not excel_path.suffix.lower() in ['.xlsx', '.xls']:
            return "错误: 文件必须是Excel格式(.xlsx或.xls)"

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        markdown_content = []

        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheets_to_process = [sheet_name]
            else:
                return f"错误: 工作表 '{sheet_name}' 不存在"
        else:
            sheets_to_process = workbook.sheetnames

        for sheet_name in sheets_to_process:
            worksheet = workbook[str(sheet_name)]

            markdown_content.append(f"# {sheet_name}")
            markdown_content.append("")

            if worksheet.max_row == 1 and worksheet.max_column == 1:
                markdown_content.append("*此工作表为空*")
                markdown_content.append("")
                continue

            rows_data = []
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        value = ""
                    row_data.append(str(value))

                if any(cell_value.strip() for cell_value in row_data):
                    rows_data.append(row_data)

            if rows_data:
                header_row = rows_data[0]
                markdown_content.append("| " + " | ".join(header_row) + " |")

                separator = "| " + " | ".join(["---"] * len(header_row)) + " |"
                markdown_content.append(separator)

                for row_data in rows_data[1:]:
                    while len(row_data) < len(header_row):
                        row_data.append("")
                    markdown_content.append("| " + " | ".join(row_data[:len(header_row)]) + " |")

            markdown_content.append("")

        result = "\n".join(markdown_content)
        return f"成功转换为Markdown！\n\n文件: {file_path}\n处理工作表: {', '.join(sheets_to_process)}\n\n{result}"

    except Exception as e:
        logger.error(f"Excel转Markdown过程中发生错误: {str(e)}")
        return f"错误: {str(e)}"
